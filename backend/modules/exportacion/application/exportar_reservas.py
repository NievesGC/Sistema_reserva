"""
Casos de uso para exportación de reservas en distintos formatos.

Implementa ExportarCSV, ExportarExcel (openpyxl) y ExportarPDF (reportlab).
Cada use case recibe una lista de reservas (dicts) y devuelve bytes o StringIO.

Columnas exportadas:
  id, servicio, fecha_desde, fecha_hasta, nombre_dueno, telefono, email,
  nombre_perro, estado, precio_total, tarifa, tramo_horario

Requisitos: 11.1, 11.2, 11.3, 11.4, 11.5
"""

from __future__ import annotations

import csv
import io
from typing import Any

from shared.infrastructure.logger import get_logger

logger = get_logger(__name__)

# Columnas a exportar (en orden)
COLUMNAS = [
    "id",
    "servicio",
    "fecha_desde",
    "fecha_hasta",
    "nombre_dueno",
    "telefono",
    "email",
    "nombre_perro",
    "estado",
    "precio_total",
    "tarifa",
    "tramo_horario",
]

# Cabeceras legibles para el usuario
CABECERAS = [
    "ID",
    "Servicio",
    "Fecha Desde",
    "Fecha Hasta",
    "Nombre Dueño",
    "Teléfono",
    "Email",
    "Nombre Perro",
    "Estado",
    "Precio Total",
    "Tarifa",
    "Tramo Horario",
]


def _extraer_fila(reserva: dict) -> list[Any]:
    """Extrae los valores de las columnas de exportación de una reserva."""
    return [reserva.get(col, "") for col in COLUMNAS]


class ExportarCSV:
    """
    Exporta una lista de reservas en formato CSV.

    Devuelve un StringIO con el contenido CSV listo para enviar como respuesta.
    """

    def execute(self, reservas: list[dict]) -> io.StringIO:
        """
        Genera el CSV con las reservas indicadas.

        Args:
            reservas: Lista de dicts con los datos de cada reserva.

        Returns:
            StringIO con el contenido CSV (UTF-8 con BOM para compatibilidad Excel).
        """
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # Cabecera
        writer.writerow(CABECERAS)

        # Filas de datos
        for reserva in reservas:
            writer.writerow(_extraer_fila(reserva))

        output.seek(0)
        logger.info("exportar_csv_ok", total_registros=len(reservas))
        return output


class ExportarExcel:
    """
    Exporta una lista de reservas en formato Excel (.xlsx) usando openpyxl.

    Devuelve bytes con el contenido del archivo Excel.
    """

    def execute(self, reservas: list[dict]) -> bytes:
        """
        Genera el archivo Excel con las reservas indicadas.

        Args:
            reservas: Lista de dicts con los datos de cada reserva.

        Returns:
            bytes con el contenido del archivo .xlsx.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise RuntimeError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"

        # Estilo de cabecera
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")

        # Escribir cabeceras con estilo
        for col_idx, cabecera in enumerate(CABECERAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=cabecera)
            cell.font = header_font
            cell.fill = header_fill

        # Escribir filas de datos
        for row_idx, reserva in enumerate(reservas, start=2):
            for col_idx, valor in enumerate(_extraer_fila(reserva), start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor is not None else "")

        # Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_length = max(
                len(str(cell.value or "")) for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        # Guardar en buffer de bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("exportar_excel_ok", total_registros=len(reservas))
        return output.getvalue()


class ExportarPDF:
    """
    Exporta una lista de reservas en formato PDF usando reportlab.

    Devuelve bytes con el contenido del archivo PDF.
    """

    def execute(self, reservas: list[dict]) -> bytes:
        """
        Genera el archivo PDF con las reservas indicadas.

        Args:
            reservas: Lista de dicts con los datos de cada reserva.

        Returns:
            bytes con el contenido del archivo .pdf.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError:
            raise RuntimeError("reportlab no está instalado. Ejecuta: pip install reportlab")

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Título
        titulo = Paragraph(
            "Listado de Reservas - Guardería Canina",
            styles["Title"],
        )
        elements.append(titulo)
        elements.append(Spacer(1, 0.5 * cm))

        # Subtítulo con total de registros
        subtitulo = Paragraph(
            f"Total de registros: {len(reservas)}",
            styles["Normal"],
        )
        elements.append(subtitulo)
        elements.append(Spacer(1, 0.5 * cm))

        # Tabla de datos
        # Usar solo columnas más relevantes para el PDF (espacio limitado)
        columnas_pdf = ["id", "servicio", "fecha_desde", "fecha_hasta", "nombre_dueno", "nombre_perro", "estado", "precio_total"]
        cabeceras_pdf = ["ID", "Servicio", "Desde", "Hasta", "Dueño", "Perro", "Estado", "Precio"]

        data = [cabeceras_pdf]
        for reserva in reservas:
            fila = [str(reserva.get(col, ""))[:20] for col in columnas_pdf]
            data.append(fila)

        tabla = Table(data, repeatRows=1)
        tabla.setStyle(TableStyle([
            # Cabecera
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            # Datos
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        elements.append(tabla)
        doc.build(elements)

        output.seek(0)
        logger.info("exportar_pdf_ok", total_registros=len(reservas))
        return output.getvalue()
